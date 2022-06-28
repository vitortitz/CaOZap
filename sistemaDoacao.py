from tkinter import CENTER, messagebox
from warnings import catch_warnings

import pandas as pd
import openpyxl as xlsx
import datetime as dt
import tkinter as tk
import pywhatkit as kit
from dateutil.relativedelta import relativedelta
from pyscreeze import center
import keyboard

from janelaCadastro import abrir_cadastro, janela_Vencimentos, janela_Aniversariantes, janela_Vencimentos7Dias

book = pd.read_excel('dados.xlsx', engine='openpyxl')


# --------------------------------------RECEBER A TABELA DO EXECEL-----------------------------------------------------------


def recebeTabela():
    tabela = xlsx.load_workbook('dados.xlsx')
    contatos = tabela.active
    return contatos


# --------------------------------------LISTA DE PAGAMENTOS-----------------------------------------------------------


lista_tipos = ["Pix", "Boleto", "Transfêrencia"]
listaContato = []


# --------------------------------------SALVAR CONTATO-----------------------------------------------------------


def salvarContato(book, entry_nome, entry_cpf, entry_nascimento, entry_numero, entry_valor, combobox_selecionarTipo,
                  entry_diaPagamento, entry_Pago):
    nome = entry_nome.get()
    cpf = entry_cpf.get()
    nascimento = entry_nascimento.get()
    numero = entry_numero.get()
    pagamento = combobox_selecionarTipo.get()
    valor = entry_valor.get()
    diaPagamento = entry_diaPagamento.get()
    pago = entry_Pago.get()
    listaContato.append((nome, cpf, nascimento, numero,
                        pagamento, valor, diaPagamento, pago))
    novo_material = pd.DataFrame(listaContato,
                                 columns=['Nome', 'CPF', 'dataNascimento', 'Número', 'métodoPagametno', 'Valor',
                                          'diaPagamento', 'Pago'])
    book = book.append(novo_material, ignore_index=True)
    book.to_excel('dados.xlsx', index=False)


# --------------------------------------VERIFICAR ANIVERSARIANTE NO DIA-----------------------------------------------------------
def verificarAniversario():
    contatos = recebeTabela()
    hoje = dt.datetime.now()
    hoje = hoje.strftime('%d/%m')
    aniversariantes = 0
    aniversariante = []
    for celula in contatos['C']:
        diaNascimento = celula.value
        if diaNascimento != "dataNascimento":
            diaNascimento = dt.datetime.strptime(diaNascimento, '%d/%m')
            diaNascimento = dt.datetime.strftime(diaNascimento, '%d/%m')
            if (diaNascimento == hoje):
                linha = celula.row
                aniversariante.append(contatos[f'A{linha}'].value)
                aniversariantes = aniversariantes + 1
    if aniversariante != []:
        return aniversariante
    else:
        aniversariante.append('Nenhum aniversariante hoje')
    return aniversariante


# --------------------------------------VERIFICAR VENCIMENTOS-----------------------------------------------------------
def verificaVencimento():
    contatos = recebeTabela()
    hoje = dt.datetime.now()
    hoje = hoje.strftime('%d/%m')
    mesatual = dt.datetime.now()
    mesatual = mesatual.strftime('%m/%y')
    vencidos = []
    for celula in contatos['G']:
        diapagamento = celula.value
        if diapagamento != "diaPagamento":
            diaPagamento = celula.value
            diaPagamento = dt.datetime.strptime(str(diaPagamento), '%d')
            mesatual = str(mesatual)
            diaPagamento = diaPagamento.strftime('%d/')
            diaPagamento = str(diaPagamento) + str(mesatual)
            linha = celula.row
            if diaPagamento < hoje and contatos[f'H{linha}'].value == 'N':
                vencidos.append(contatos[f'A{linha}'].value)
    if vencidos != []:
        return vencidos
    else:
        vencidos.append('Nenhum vencimento acima de 3 dias')
    return vencidos


# --------------------------------------VERIFICAR VENCIMENTOS ACIMA DE 7 DIAS-----------------------------------------------------------
def vencimento3Dias():
    contatos = recebeTabela()
    hoje = dt.datetime.now().strftime('%d/%m/%y')
    vencidos = []
    for celula in contatos['G']:
        diapagamento = celula.value
        if diapagamento != "diaPagamento":
            diaPagamento = celula.value
            linha = celula.row
            diatual = dt.datetime.strptime(hoje, '%d/%m/%y')
            dias = abs(((diatual) - diaPagamento).days)
            if dias > 3 and contatos[f'H{linha}'].value == 'N':
                vencidos.append(contatos[f'A{linha}'].value)
    if vencidos != []:
        return vencidos
    else:
        vencidos.append('Nenhum vencimento acima de 7 dias')
    return vencidos


# --------------------------------------VERIFICAR VENCIMENTOS ACIMA DE 7 DIAS-----------------------------------------------------------
def vencimentos7Dias():
    contatos = recebeTabela()
    hoje = dt.datetime.now().strftime('%d/%m/%y')
    vencidos = []
    for celula in contatos['G']:
        diapagamento = celula.value
        if diapagamento != "diaPagamento":
            diaPagamento = celula.value
            linha = celula.row
            diatual = dt.datetime.strptime(hoje, '%d/%m/%y')
            dias = abs(((diatual) - diaPagamento).days)
            if dias > 7 and contatos[f'H{linha}'].value == 'N':
                vencidos.append(contatos[f'A{linha}'].value)
    if vencidos != []:
        return vencidos
    else:
        vencidos.append('Nenhum vencimento acima de 7 dias')
    return vencidos


# --------------------------------------ENVIAR MENSAGEM AOS VENCIDOS DE 3 DIAS-----------------------------------------------------------

def enviarMensagemVencidos():
    contatos = recebeTabela()
    vencidos = vencimento3Dias()
    numeros = []
    for celula in contatos['A']:
        nome = celula.value
        if nome != "Nome":
            for vencido in vencidos:
                if nome == vencido:
                    linha = celula.row
                    numeros.append(contatos[f'D{linha}'].value)
    try:
        for numero in numeros:
            kit.sendwhats_image(
                f'+{numero}', r"cobranca.jpeg", "Bom dia", 25, True, 5)
        messagebox.showinfo("Finalizado", "Todas as mensagens enviadas")
    except KeyboardInterrupt:
        messagebox.showinfo(
            "Finalizado", "O envio de mensagem foi interrompido!")
# --------------------------------------ENVIAR MENSAGEM AOS VENCIDOS DE 7 DIAS-----------------------------------------------------------


def enviarMensagemVencidos7Dias():
    contatos = recebeTabela()
    vencidos = verificaVencimento()
    numeros = []
    for celula in contatos['A']:
        nome = celula.value
        if nome != "Nome":
            for vencido in vencidos:
                if nome == vencido:
                    linha = celula.row
                    numeros.append(contatos[f'D{linha}'].value)
        while True:
            try:
                for numero in numeros:
                    if keyboard.is_pressed('q'):
                        print('PAROU')
                    kit.sendwhats_image(
                        f'+{numero}', r"cobranca2.jpeg", "Bom dia", 25, True, 5)
                    messagebox.showinfo(
                        "Finalizado", "Todas as mensagens enviadas")
            except:
                break

# --------------------------------------ENVIAR MENSAGEM DE ANIVERSARIO-----------------------------------------------------------


def enviarMensagemAniversario():
    contatos = recebeTabela()
    aniversariantes = verificarAniversario()
    if aniversariantes != ['Nenhum aniversariante hoje']:
        numeros = []
        for celula in contatos['A']:
            nome = celula.value
            if nome != "Nome":
                for aniversariante in aniversariantes:
                    if nome == aniversariante:
                        linha = celula.row
                        numeros.append(contatos[f'D{linha}'].value)
        for numero in numeros:
            kit.sendwhats_image(f'+{numero}', r"aniversario.jpeg", "", 25, True,
                                5)
    messagebox.showinfo("Finalizado", "Todas as mensagens enviadas")


def trocaPago():
    tabela = xlsx.load_workbook('dados.xlsx')
    contatos = tabela.active
    hoje = dt.datetime.now().strftime('%d')
    hoje = hoje.lstrip('0')
    for celula in contatos['G']:
        diapagamento = celula.value
        if diapagamento != "diaPagamento":
            diaPagamento = celula.value
            linha = celula.row
            if str(diaPagamento) == str(hoje):
                contatos[f'H{linha}'] = 'N'
    tabela.save("dados.xlsx")


def trocaMes():
    tabela = xlsx.load_workbook('dados.xlsx')
    contatos = tabela.active
    hoje = dt.datetime.now().strftime('%d/%m/%y')
    for celula in contatos['G']:
        diapagamento = celula.value
        if diapagamento != "diaPagamento":
            diaPagamento = celula.value
            linha = celula.row
            diatual = dt.datetime.strptime(hoje, '%d/%m/%y')
            dias = abs(((diatual) - diaPagamento).days)
            if dias > 30:
                contatos[f'G{linha}'] = diaPagamento + relativedelta(months=1)
    tabela.save("dados.xlsx")


# --------------------------------------JANELA DE CADASTRO-----------------------------------------------------------
def cadastros():
    abrir_cadastro(lista_tipos, salvarContato, book)


# --------------------------------------JANELA DE VENCIMENTOS-----------------------------------------------------------
def vencimento():
    janela_Vencimentos(vencimento3Dias, enviarMensagemVencidos)


def vencimento7Dias():
    janela_Vencimentos7Dias(vencimentos7Dias, enviarMensagemVencidos7Dias)


# --------------------------------------JANELA DE ANIVERSARIANTE-----------------------------------------------------------
def aniversariante():
    janela_Aniversariantes(verificarAniversario, enviarMensagemAniversario)


# --------------------------------------JANELA PRINCIAPAL-----------------------------------------------------------

janela = tk.Tk()
janela.title('Menu de Opções')
altura = 140
largura = 250
largura_tela = janela.winfo_screenwidth()
altura_tela = janela.winfo_screenheight()
posx = largura_tela/2 - largura/2
posy = altura_tela/2 - altura/2
janela.geometry('%dx%d+%d+%d' % (largura, altura, posx, posy))

botaoCadastrarDoador = tk.Button(
    janela, text="Cadastrar Doador", command=cadastros)
botaoCadastrarDoador.place(relx=0.5, rely=0.15, anchor=CENTER)
botaoVerificaAniversariante = tk.Button(
    janela, text="Verificar Aniversariante", command=(aniversariante))
botaoVerificaAniversariante.place(relx=0.5, rely=0.4, anchor=CENTER)
botaoVerificaVencimento = tk.Button(
    janela, text="Verificar Vencimentos 3 Dias", command=(vencimento))
botaoVerificaVencimento.place(relx=0.5, rely=0.65, anchor=CENTER)
botaoVerificaVencimento7Dias = tk.Button(
    janela, text="Verificar Vencimentos 7 Dias", command=(vencimento7Dias))
botaoVerificaVencimento7Dias.place(relx=0.5, rely=0.9, anchor=CENTER)
trocaPago()
trocaMes()
janela.mainloop()
